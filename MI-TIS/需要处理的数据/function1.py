# %%
import openpyxl
from scipy import optimize
import numpy as np


# %%
# 插入字符串
def insert_string(string, need, num):
    a = list(string)
    a.insert(num, need)
    b = ''.join(a)
    return b


# 直线拟合函数
def f_1(x, A, B):
    return A * x + B


# 非线性拟合函数
def func(x, a, b, c):
    return a*x*x+b*x+c


# 计算R2的值
def caculate_R2(data3, data1, A1=[], name='', A2=0, B2=0):
    if 'MI' in name:
        RSS = np.dot((data1 - f_1(data1, A2, B2)), (data3 - f_1(data1, A2, B2)))
    else:
        RSS = np.dot((data1 - func(data1, *A1)), (data3 - func(data1, *A1)))
    ymean = np.mean(data3)
    TSS = np.dot((data3 - ymean), (data3 - ymean))
    R2 = 1 - RSS / TSS
    return R2


# %%
# 创建excel表
def creat_excel(a, b, file_string, sheet_name, count):
    if count <= 0:
        wb = openpyxl.Workbook()                  # 创建excel表
        ws = wb.create_sheet(sheet_name, count)   # 创建新的sheet表
        del wb['Sheet']                           # 删除sheet

        # 表
    else:
        wb = openpyxl.load_workbook(file_string)  # 读取excel文件
        ws = wb.create_sheet(sheet_name, count)   # 创建新的sheet表

    row = 2                                                                     # 行数
    column = 1                                                                  # 列数
    bold = openpyxl.styles.Font(bold=True)                                      # 设置字体加粗
    center = openpyxl.styles.Alignment(horizontal='center', vertical='center')  # 设置垂直居中和水平居中

    ws['A1'] = 'powerlever'      # A1表格输入powerlever
    ws['A1'].font = bold         # 设置字体加粗
    ws['A1'].alignment = center  # 设置垂直居中和水平居中
    ws['B1'] = 'MI'              # B1表格输入MI
    ws['B1'].font = bold         # 设置字体加粗
    ws['B1'].alignment = center  # 设置垂直居中和水平居中
    ws['C1'] = 'TIS'             # C1表格输入TIS
    ws['C1'].font = bold         # 设置字体加粗
    ws['C1'].alignment = center  # 设置垂直居中和水平居中

    # 循环输入powerlever值
    for cost in a[0]:
        ws.cell(row, column).value = cost
        row += 1

    row = 2  # 重置行数

    # 循环输入MI值
    for item in a[1]:
        ws.cell(row, column + 1).value = item
        row += 1

    row = 2  # 重置行数

    # 循环输入TIS值
    for item in b[1]:
        ws.cell(row, column + 2).value = item
        row += 1

    ws.cell(1, 5).value = 'R2_MI=' + str(a[2])    # 输入MI的拟合度R2的值
    ws.cell(2, 5).value = 'R2_TIS=' + str(b[2])  # 输入TIS的拟合度R2的值

    wb.save(file_string)                           # excel保存路径


# %%
# 创建图像以及保存图像
def creat_image(data3, data1, y):

    # 判断使用哪种函数
    if 'MI' in y:
        A1, B1 = optimize.curve_fit(f_1, data3, data1)[0]     # 进行线性拟合
        x1 = np.arange(0, 101, 1)                             # 拟合的范围以及每次点的间距
        y1 = A1 * x1 + B1                                     # 拟合后y的值
        R2 = caculate_R2(data3, data1, name=y, A2=A1, B2=B1)  # 计算R2的值

    else:
        A1, B1 = optimize.curve_fit(func, data3, data1)       # 进行非线性拟合
        # 提取需要的相应变量
        a = A1[0]
        b = A1[1]
        c = A1[2]
        x1 = np.arange(0, 101, 1)                             # 拟合的范围以及每次点的间距
        y1 = func(x1, a, b, c)                                # 拟合后y的值
        R2 = caculate_R2(data3, data1, A1=A1, name=y)         # 计算R2的值

    return x1, y1, R2                                         # 返回保存数据
