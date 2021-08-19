import os
import xlrd
import openpyxl


# 保存相应文件夹下一级文件夹名称
def save_folders(a):
	b = []                                                  # 创建空列表
	for root, dirs, file in os.walk(a, topdown=False):
		b = dirs                                            # 保存文件夹名称
	return b                                                # 返回文件夹列表


# 保存相应文件夹下一级文件名称
def save_file(a):
	c = []                                                  # 创建空列表
	for root, dirs, file in os.walk(a, topdown=False):
		c = file                                            # 保存文件名称
		break                                               # 跳出循环
	return c                                                # 返回文件名称


# 对ROI或RES进行排序
def sort(file, folder):
	list_number = []                                        # 文件夹中的数字列表
	list = []                                               # 最终返回的列表
	# 判断需要使用哪个部分
	if 'C Mode' in folder:
		for item in file:
			item1 = item.replace('ROI', '')                 # 删除文件夹名称中的ROI
			list_number.append(int(item1))                  # 提取数字并增加到列表中
		list_number.sort()                                  # 对数字进行排序
		for number in list_number:
			file_name = 'ROI' + str(number)                 # 文件夹名称还原
			list.append(file_name)                          # 增加文件夹名称
		return list                                         # 返回最终排序好的文件夹名称列表
	elif 'B Mode' in folder:
		for item in file:
			item1 = item.replace('RES', '')                 # 删除文件夹名称中的RES
			list_number.append(int(item1))                  # 提取数字并增加到列表中
		list_number.sort()                                  # 对数字进行排序
		for number in list_number:
			file_name = 'RES' + str(number)                 # 文件夹名称还原
			list.append(file_name)                          # 增加文件夹名称
		return list                                         # 返回最终排序好的文件夹名称列表
	else:
		return file                                         # 返回文件夹名称


# 对文件名进行排序
def sort_file(file):
	list_number = []                                        # 文件夹中的数字列表
	list = []                                               # 最终返回的列表
	string1 = file[0][::-1]                                 # 对文件名称进行颠倒
	string2 = string1.split('_', 1)[1]                      # 提取文件名称需要的字符串
	string3 = string2[::-1]                                 # 最终需要的字符串

	for item in file:
		item1 = item.replace('%.xls', '')                   # 删除文件名称中的%.xls
		number_file = item1.split('_')[-1]                  # 提取文件名称中的最后数字
		list_number.append(int(number_file))                # 将数字增加到数字列表中

	list_number.sort()                                      # 对数字进行排序
	for number in list_number:
		file_name = string3 + '_' + str(number) + '%.xls'   # 将文件名称进行复原
		list.append(file_name)                              # 增加文件名称
	return list                                             # 返回最终排序好的文件名称列表


# 提取excel文件中MI/TIS的值
def extract_MI_TIS(path):
	wb = xlrd.open_workbook(path)                           # 打开相应的excle表
	ws = wb.sheet_by_name('Output')                         # 需要提取的sheet表名称
	MI = ws.cell(54, 2).value                               # 提取MI数据
	TIS = ws.cell(55, 2).value                              # 提取TIS数据
	return MI, TIS                                          # 返回MI,TIS的值


# 创建excel表
def creat_excel(data1, data2, data3, count, sheet_name, file_string):
	# 判断创建新的excle表还是创建新的sheet表
	if count <= 0:
		wb = openpyxl.Workbook()                            # 创建excel表
		ws = wb.create_sheet(sheet_name, count)             # 创建新的sheet表
		del wb['Sheet']                                     # 删除sheet

	# 表
	else:
		wb = openpyxl.load_workbook(file_string)            # 读取excel文件
		ws = wb.create_sheet(sheet_name, count)             # 创建新的sheet表

	row = 3                                                 # 行数
	column = 1                                              # 列数
	bold = openpyxl.styles.Font(bold=True)                  # 设置字体加粗
	center = openpyxl.styles.Alignment(horizontal='center', vertical='center')  # 设置垂直居中和水平居中

	ws['A1'] = 'powerlevel'                                 # A1表格输入powerlevel
	ws['A1'].font = bold                                    # 设置字体加粗
	ws['A1'].alignment = center                             # 设置垂直居中和水平居中
	ws['B1'] = 'MI'                                         # B1表格输入MI
	ws['B1'].font = bold                                    # 设置字体加粗
	ws['B1'].alignment = center                             # 设置垂直居中和水平居中
	ws['C1'] = 'TIS'                                        # C1表格输入TIS
	ws['C1'].font = bold                                    # 设置字体加粗
	ws['C1'].alignment = center                             # 设置垂直居中和水平居中
	ws['A2'] = 0
	ws['B2'] = 0
	ws['C2'] = 0

	# 循环输入powerlever值
	for cost in data3:
		ws.cell(row, column).value = cost
		row += 1

	row = 3                                                 # 重置行数

	# 循环输入MI值
	for item in data1:
		ws.cell(row, column + 1).value = item
		row += 1

	row = 3                                                 # 重置行数

	# 循环输入TIS值
	for item in data2:
		ws.cell(row, column + 2).value = item
		row += 1

	wb.save(file_string)                                     # excel保存路径


# 创建文件夹
def creat_folder(path, folder):
	folder_path = path + folder                              # 文件夹路径
	a = os.path.exists(folder_path)                          # 返回文件夹是否存在
	if a:
		a = 1
	else:
		os.mkdir(folder_path)                                # 不存在文件夹创建新的文件夹
